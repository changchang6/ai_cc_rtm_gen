#!/usr/bin/env python3
"""
TBUS RTM Generator
Generate RTM file with Checker List and DV Testcase List for TBUS module.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy
import os

# Define Checker List based on LRS and TP requirements
CHECKERS = [
    {
        "chk_id": "CHK_001",
        "chk_name": "clk_freq_check",
        "chk_description": """1.检查频率值是否正确：连续采集clk_i时钟上升沿和下降沿并计算实际频率/周期，判断是否处于目标频率100MHz的允许误差范围内（±1%）；
2.检查时钟稳定性：通过连续采样周期，对比相邻两个周期的频率偏差是否在3%内。""",
        "note": ""
    },
    {
        "chk_id": "CHK_002",
        "chk_name": "reset_check",
        "chk_description": """1.检查rst_ni异步复位、同步释放机制：当rst_ni=0时，所有寄存器立即复位；rst_ni=1释放后，需经过同步处理才能生效；
2.检查复位后状态机状态：state应为IDLE状态（state==IDLE）；
3.检查复位后寄存器默认值：CTRL.EN=0，CTRL.LANE_MODE=2'b00（1-bit模式），STATUS=0，LAST_ERR=0；
4.检查协议上下文清空：rx_shift_reg=0，tx_shift_reg=0，opcode_reg=0，addr_reg=0，wdata_reg=0，status_code=0。""",
        "note": ""
    },
    {
        "chk_id": "CHK_003",
        "chk_name": "csr_access_check",
        "chk_description": """1.检查VERSION寄存器读取：通过RD_CSR(8'h11)访问VERSION，返回值应符合设计预期；
2.检查CTRL寄存器读写：通过WR_CSR(8'h10)写入CTRL.EN/LANE_MODE/SOFT_RST，再通过RD_CSR(8'h11)读回验证；
3.检查STATUS寄存器读取：STATUS[0]反映busy状态，STATUS[7:0]在写类命令后返回status_code；
4.检查LAST_ERR寄存器：错误发生时LAST_ERR应更新为对应错误码（0x01~0x08）。""",
        "note": ""
    },
    {
        "chk_id": "CHK_004",
        "chk_name": "test_mode_check",
        "chk_description": """1.检查test_mode_i=1且CTRL.EN=1时，模块允许执行功能命令；
2.检查test_mode_i=0时，模块拒绝执行命令并返回STS_NOT_IN_TEST(0x05)；
3.检查CTRL.EN=0时，模块返回STS_DISABLED(0x04)；
4.检查半双工模式：请求阶段(pdo_oe_o=0)接收数据，turnaround周期后响应阶段(pdo_oe_o=1)输出数据。""",
        "note": ""
    },
    {
        "chk_id": "CHK_005",
        "chk_name": "lane_mode_check",
        "chk_description": """1.检查1-bit模式(LANE_MODE=2'b00)：仅使用pdi_i[0]/pdo_o[0]传输，MSB first；
2.检查4-bit模式(LANE_MODE=2'b01)：使用pdi_i[3:0]/pdo_o[3:0]传输，高nibble优先发送；
3.检查非法LANE_MODE(2'b10/2'b11)：应返回错误或保持原配置；
4.检查Burst/AXI功能不可用：尝试触发Burst或AXI操作应返回错误或不响应。""",
        "note": ""
    },
    {
        "chk_id": "CHK_006",
        "chk_name": "data_interface_check",
        "chk_description": """1.检查pcs_n_i帧边界：pcs_n_i=0开启事务，pcs_n_i=1关闭事务；
2.检查pdi_i输入数据：在pcs_n_i=0期间有效，按LANE_MODE模式接收；
3.检查pdo_o输出数据：在响应阶段pdo_oe_o=1时有效，按LANE_MODE模式发送；
4.检查pdo_oe_o方向控制：请求阶段=0（输入），turnaround周期后=1（输出）；
5.检查MSB first传输顺序：opcode、addr、data均按MSB first顺序传输。""",
        "note": ""
    },
    {
        "chk_id": "CHK_007",
        "chk_name": "protocol_timing_check",
        "chk_description": """1.检查请求/响应阶段分离：请求阶段pdo_oe_o=0接收pdi_i，响应阶段pdo_oe_o=1输出pdo_o；
2.检查turnaround周期：请求阶段结束后固定1个clk_i周期turnaround，期间pdo_oe_o=0；
3.检查四种命令帧格式正确解析：
   - WR_CSR(8'h10): opcode(8bit) + addr(8bit) + wdata(32bit)
   - RD_CSR(8'h11): opcode(8bit) + addr(8bit)
   - AHB_WR32(8'h20): opcode(8bit) + addr(32bit) + wdata(32bit)
   - AHB_RD32(8'h21): opcode(8bit) + addr(32bit)；
4.检查响应帧格式：status(8bit) [+ rdata(32bit)]，读类命令包含rdata。""",
        "note": ""
    },
    {
        "chk_id": "CHK_008",
        "chk_name": "opcode_decode_check",
        "chk_description": """1.检查合法opcode正确译码：8'h10→WR_CSR，8'h11→RD_CSR，8'h20→AHB_WR32，8'h21→AHB_RD32；
2.检查非法opcode(如8'h00/8'hFF)错误处理：返回STS_BAD_OPCODE(0x01)；
3.检查前端帧长确定：收到8bit opcode后，根据opcode类型确定期望帧长；
4.检查opcode与后续数据一致性：opcode类型与addr/wdata长度匹配。""",
        "note": ""
    },
    {
        "chk_id": "CHK_009",
        "chk_name": "ctrl_config_check",
        "chk_description": """1.检查CTRL.EN使能/禁用：EN=1时功能命令可执行，EN=0时返回STS_DISABLED(0x04)；
2.检查CTRL.LANE_MODE配置：00→1-bit模式，01→4-bit模式，10/11→非法或忽略；
3.检查CTRL.SOFT_RST软复位：写1触发协议上下文清零并恢复默认lane模式(1-bit)；
4.检查配置生效时序：配置写入后下一周期生效。""",
        "note": ""
    },
    {
        "chk_id": "CHK_010",
        "chk_name": "status_polling_check",
        "chk_description": """1.检查STATUS[0] busy位：命令执行期间busy=1，完成后busy=0；
2.检查STATUS[7:0]状态码：命令完成后返回对应的status_code；
3.检查LAST_ERR寄存器：错误发生时更新为对应错误码，保持直到下次错误或复位；
4.检查无独立中断输出：MVP版本通过轮询STATUS/LAST_ERR获取状态，无irq信号输出。""",
        "note": ""
    },
    {
        "chk_id": "CHK_011",
        "chk_name": "error_detection_check",
        "chk_description": """1.检查BAD_OPCODE(0x01)：非法opcode输入时返回status_code=8'h01；
2.检查BAD_REG(0x02)：访问非法CSR地址时返回status_code=8'h02；
3.检查ALIGN_ERR(0x03)：AHB地址非4Byte对齐时返回status_code=8'h03；
4.检查DISABLED(0x04)：CTRL.EN=0时发送命令返回status_code=8'h04；
5.检查NOT_IN_TEST(0x05)：test_mode_i=0时发送命令返回status_code=8'h05；
6.检查FRAME_ERR(0x06)：pcs_n_i提前拉高（帧错误）返回status_code=8'h06；
7.检查AHB_ERR(0x07)：hresp_i=1时返回status_code=8'h07；
8.检查TIMEOUT(0x08)：AHB访问超时时返回status_code=8'h08；
9.检查错误收敛：在发起AHB访问前完成前置错误检测。""",
        "note": ""
    },
    {
        "chk_id": "CHK_012",
        "chk_name": "low_power_check",
        "chk_description": """1.检查空闲态功耗优化：test_mode_i=0或无有效任务时，接收/发送移位寄存器不翻转；
2.检查超时计数器：仅在RX/TX/WAIT_AHB相关状态翻转；
3.检查AHB输出IDLE：test_mode_i=0时htrans_o=IDLE，hwrite_o=0；
4.检查关键寄存器翻转：仅在工作状态翻转，空闲态保持。""",
        "note": ""
    },
    {
        "chk_id": "CHK_013",
        "chk_name": "performance_check",
        "chk_description": """1.检查目标频率：clk_i稳定工作在100MHz；
2.检查端到端延时：从pcs_n_i拉低到响应完成的总周期数约23~24 cycles；
3.检查吞吐率：1-bit模式约12.5MB/s，4-bit模式约50MB/s；
4.检查AHB访问约束：hsize_o固定为WORD(32-bit)，地址4Byte对齐。""",
        "note": ""
    },
    {
        "chk_id": "CHK_014",
        "chk_name": "dfx_observability_check",
        "chk_description": """1.检查状态机状态可观测：state信号可采集；
2.检查关键寄存器可观测：opcode、addr、wdata、rdata、status_code可采集；
3.检查rx/tx计数可观测：接收/发送数据计数器可访问；
4.检查接口信号可观测：pcs_n_i、pdi_i、pdo_o、pdo_oe_o可采集。""",
        "note": ""
    },
    {
        "chk_id": "CHK_015",
        "chk_name": "memory_map_check",
        "chk_description": """1.检查CSR通过协议访问：VERSION/CTRL/STATUS/LAST_ERR通过WR_CSR/RD_CSR访问，不进入SoC AHB memory map；
2.检查AHB Master地址空间：32bit地址空间，word访问粒度；
3.检查AHB Master访问目标：可达地址范围由SoC顶层约束。""",
        "note": ""
    },
    {
        "chk_id": "CHK_016",
        "chk_name": "ahb_interface_check",
        "chk_description": """1.检查haddr_o输出：32bit地址，4Byte对齐；
2.检查hwrite_o输出：1=写，0=读；
3.检查htrans_o输出：仅IDLE(2'b00)和NONSEQ(2'b10)；
4.检查hsize_o输出：固定为WORD(3'b010)；
5.检查hburst_o输出：固定为SINGLE(3'b000)；
6.检查hwdata_o输出：写数据在数据相有效；
7.检查hrdata_i输入：读返回数据在hready_i=1时有效；
8.检查hready_i输入：等待状态指示；
9.检查hresp_i输入：错误响应(OKAY=0, ERROR=1)。""",
        "note": ""
    }
]

# Define Testcase List based on LRS and TP requirements
TESTCASES = [
    {
        "tc_id": "TC_001",
        "tc_name": "tbus_clk_freq_test",
        "tc_description": """配置条件：
1.配置clk_i时钟频率为100MHz；
2.配置test_mode_i=1，CTRL.EN=1；
3.配置LANE_MODE为1-bit或4-bit模式。

输入激励：
1.连续发送AHB_RD32(8'h21)和AHB_WR32(8'h20)命令序列，覆盖1-bit和4-bit两种lane模式；
2.每种模式执行不少于100次命令事务。

期望结果：
1.所有命令在100MHz频率下正确执行，协议采样、状态机控制和AHB-Lite主接口同域工作正常；
2.无时序违例，频率偏差在±1%范围内。

coverage check点：
直接用例覆盖，不收功能覆盖率""",
        "note": ""
    },
    {
        "tc_id": "TC_002",
        "tc_name": "tbus_reset_test",
        "tc_description": """配置条件：
1.初始状态rst_ni为低，模块处于复位状态。

输入激励：
1.释放rst_ni，等待同步释放完成；
2.通过RD_CSR(8'h11)读取CTRL寄存器，验证默认值；
3.检查STATUS和LAST_ERR寄存器清零；
4.写入非默认值到CTRL后再次复位，验证复位后恢复默认值。

期望结果：
1.复位后状态机处于IDLE状态；
2.CTRL.EN=0，CTRL.LANE_MODE=2'b00（1-bit模式）；
3.STATUS=0，LAST_ERR=0；
4.协议上下文（rx_shift_reg、tx_shift_reg、opcode_reg等）已清空。

coverage check点：
直接用例覆盖，不收功能覆盖率""",
        "note": ""
    },
    {
        "tc_id": "TC_003",
        "tc_name": "tbus_csr_access_test",
        "tc_description": """配置条件：
1.模块使能(test_mode_i=1, CTRL.EN=1)；
2.配置LANE_MODE为1-bit或4-bit模式。

输入激励：
1.通过RD_CSR(8'h11)读取VERSION寄存器，验证版本号符合预期；
2.通过WR_CSR(8'h10)写入CTRL.EN=1，再通过RD_CSR读回验证；
3.通过WR_CSR配置LANE_MODE=00/01，读回验证；
4.触发错误条件后读取STATUS和LAST_ERR寄存器。

期望结果：
1.VERSION寄存器返回预期版本值；
2.CTRL寄存器读写正确，配置生效；
3.STATUS[0]正确反映busy状态；
4.LAST_ERR在错误发生时更新为对应错误码。

coverage check点：
直接用例覆盖，不收功能覆盖率""",
        "note": ""
    },
    {
        "tc_id": "TC_004",
        "tc_name": "tbus_test_mode_test",
        "tc_description": """配置条件：
1.初始test_mode_i=0，CTRL.EN=0。

输入激励：
1.配置test_mode_i=1，通过WR_CSR(8'h10)设置CTRL.EN=1；
2.发送WR_CSR命令配置lane模式；
3.发送AHB_WR32(8'h20)/AHB_RD32(8'h21)命令进行单次读写操作；
4.验证半双工请求/响应模式：检查pdo_oe_o在请求阶段为0，响应阶段为1；
5.清除CTRL.EN=0，再次发送命令验证返回STS_DISABLED。

期望结果：
1.模块正确执行功能命令；
2.协议采用半双工模式，请求/响应阶段正确分离；
3.CTRL.EN=0时返回STS_DISABLED(0x04)；
4.test_mode_i=0时返回STS_NOT_IN_TEST(0x05)。

coverage check点：
直接用例覆盖，不收功能覆盖率""",
        "note": ""
    },
    {
        "tc_id": "TC_005",
        "tc_name": "tbus_lane_mode_test",
        "tc_description": """配置条件：
1.模块使能(test_mode_i=1, CTRL.EN=1)。

输入激励：
1.配置LANE_MODE=2'b00验证1-bit模式：发送AHB_RD32命令，检查仅pdi_i[0]/pdo_o[0]有效；
2.配置LANE_MODE=2'b01验证4-bit模式：发送相同命令，检查pdi_i[3:0]/pdo_o[3:0]按高nibble优先发送；
3.尝试配置非法LANE_MODE值(2'b10/2'b11)验证错误处理或忽略；
4.验证Burst和AXI功能不可用：尝试触发Burst命令或AXI操作。

期望结果：
1.1-bit模式正常工作，仅bit0传输数据，MSB first；
2.4-bit模式正常工作，高nibble优先发送；
3.非法LANE_MODE配置返回错误或保持原配置；
4.Burst和AXI功能确认不可用。

coverage check点：
直接用例覆盖，不收功能覆盖率""",
        "note": ""
    },
    {
        "tc_id": "TC_006",
        "tc_name": "tbus_data_interface_test",
        "tc_description": """配置条件：
1.模块使能，配置1-bit和4-bit lane模式；
2.ATE模拟器准备发送/接收数据。

输入激励：
1.在1-bit模式下发送WR_CSR(8'h10)/AHB_WR32(8'h20)命令，验证pdi_i[0]/pdo_o[0]数据传输；
2.在4-bit模式下发送相同命令，验证pdi_i[3:0]/pdo_o[3:0]按高nibble优先发送；
3.验证pcs_n_i帧边界：拉低开启事务，拉高关闭事务；
4.检查pdo_oe_o方向控制：请求阶段=0，turnaround后=1。

期望结果：
1.数据接口按lane模式正确传输；
2.帧边界清晰，pcs_n_i控制正确；
3.MSB first传输顺序正确；
4.pdo_oe_o方向控制正确。

coverage check点：
直接用例覆盖，不收功能覆盖率""",
        "note": ""
    },
    {
        "tc_id": "TC_007",
        "tc_name": "tbus_protocol_timing_test",
        "tc_description": """配置条件：
1.模块使能，配置不同lane模式；
2.配置AHB总线响应正常。

输入激励：
1.发送WR_CSR(8'h10)命令验证请求阶段：pdi_i输入opcode+addr+wdata；
2.验证turnaround周期：请求阶段结束后固定1个clk_i周期pdo_oe_o=0；
3.检查响应阶段pdo_oe_o=1输出status+rdata；
4.验证四种命令帧格式：WR_CSR(8'h10)、RD_CSR(8'h11)、AHB_WR32(8'h20)、AHB_RD32(8'h21)。

期望结果：
1.请求/响应阶段正确分离；
2.turnaround固定1周期；
3.协议字段MSB first传输；
4.四种帧格式正确处理，帧长度符合规范。

coverage check点：
直接用例覆盖，不收功能覆盖率""",
        "note": ""
    },
    {
        "tc_id": "TC_008",
        "tc_name": "tbus_opcode_decode_test",
        "tc_description": """配置条件：
1.模块使能，AHB总线空闲；
2.配置LANE_MODE为1-bit或4-bit模式。

输入激励：
1.发送opcode=8'h10执行WR_CSR命令，验证正确译码；
2.发送opcode=8'h11执行RD_CSR命令，验证正确译码；
3.发送opcode=8'h20执行AHB_WR32命令，验证正确译码；
4.发送opcode=8'h21执行AHB_RD32命令，验证正确译码；
5.发送非法opcode(如8'h00/8'hFF)验证返回STS_BAD_OPCODE(0x01)。

期望结果：
1.四种opcode正确译码并执行对应操作；
2.非法opcode返回status_code=8'h01；
3.前端在收到8bit opcode后正确确定期望帧长。

coverage check点：
直接用例覆盖，不收功能覆盖率""",
        "note": ""
    },
    {
        "tc_id": "TC_009",
        "tc_name": "tbus_ctrl_config_test",
        "tc_description": """配置条件：
1.模块初始状态，test_mode_i=1。

输入激励：
1.写CTRL.EN=1使能模块，发送AHB_RD32命令验证功能命令可执行；
2.写CTRL.EN=0禁用模块，发送命令验证返回STS_DISABLED(0x04)；
3.配置LANE_MODE=2'b00验证1-bit模式生效；
4.配置LANE_MODE=2'b01验证4-bit模式生效；
5.写CTRL.SOFT_RST=1触发软复位，验证协议上下文清零并恢复默认lane模式(1-bit)。

期望结果：
1.CTRL.EN配置正确生效，使能/禁用状态切换正确；
2.LANE_MODE配置正确生效；
3.SOFT_RST触发软复位，协议上下文清零，恢复默认配置。

coverage check点：
直接用例覆盖，不收功能覆盖率""",
        "note": ""
    },
    {
        "tc_id": "TC_010",
        "tc_name": "tbus_status_polling_test",
        "tc_description": """配置条件：
1.模块使能，执行各类命令。

输入激励：
1.触发各类错误条件(非法opcode、帧错误、AHB错误等)；
2.通过RD_CSR(8'h11)读取STATUS寄存器检查busy位和status_code；
3.读取LAST_ERR寄存器获取最近错误码；
4.验证无独立中断输出信号。

期望结果：
1.命令执行期间STATUS[0]=1，完成后=0；
2.STATUS[7:0]正确反映命令执行状态；
3.LAST_ERR在错误发生时更新为对应错误码；
4.MVP版本无中断输出，ATE轮询方式正确获取状态。

coverage check点：
直接用例覆盖，不收功能覆盖率""",
        "note": ""
    },
    {
        "tc_id": "TC_011",
        "tc_name": "tbus_error_detection_test",
        "tc_description": """配置条件：
1.模块使能，AHB总线正常；
2.配置各种错误触发条件。

输入激励：
1.发送非法opcode(8'h00/8'hFF)验证返回STS_BAD_OPCODE(0x01)；
2.访问非法CSR地址(如8'hFF)验证返回STS_BAD_REG(0x02)；
3.发送非4Byte对齐地址(如32'h0000_0001)验证返回STS_ALIGN_ERR(0x03)；
4.CTRL.EN=0时发送命令验证返回STS_DISABLED(0x04)；
5.test_mode_i=0时发送命令验证返回STS_NOT_IN_TEST(0x05)；
6.提前拉高pcs_n_i验证返回STS_FRAME_ERR(0x06)；
7.触发hresp_i=1验证返回STS_AHB_ERR(0x07)；
8.模拟AHB超时验证返回STS_TIMEOUT(0x08)。

期望结果：
1.各类错误正确检测；
2.在发起AHB访问前完成前置错误收敛；
3.返回确定性状态码；
4.LAST_ERR更新正确。

coverage check点：
直接用例覆盖，不收功能覆盖率""",
        "note": ""
    },
    {
        "tc_id": "TC_012",
        "tc_name": "tbus_low_power_test",
        "tc_description": """配置条件：
1.模块处于空闲态(test_mode_i=0或无有效任务)。

输入激励：
1.保持pcs_n_i=1验证接收/发送移位寄存器不翻转；
2.模拟AHB等待态验证超时计数器仅在此状态工作；
3.test_mode_i=0时验证htrans_o=IDLE，hwrite_o=0；
4.切换到工作态验证关键寄存器正常翻转。

期望结果：
1.空闲态无效翻转最小化；
2.低功耗设计目标满足；
3.AHB输出保持IDLE状态。

coverage check点：
直接用例覆盖，不收功能覆盖率""",
        "note": ""
    },
    {
        "tc_id": "TC_013",
        "tc_name": "tbus_performance_test",
        "tc_description": """配置条件：
1.配置clk_i=100MHz；
2.AHB总线响应正常(hready_i=1，hresp_i=0)。

输入激励：
1.在1-bit模式下连续执行AHB_RD32/AHB_WR32命令，计算端到端延时和吞吐率；
2.在4-bit模式下执行相同测试；
3.验证原始链路线速：1-bit模式12.5MB/s，4-bit模式50MB/s；
4.验证AHB接口32bit字访问和4Byte对齐约束。

期望结果：
1.目标频率100MHz稳定工作；
2.端到端延时约23~24 cycles；
3.吞吐率达标：1-bit模式≥12.5MB/s，4-bit模式≥50MB/s；
4.AHB访问约束满足：hsize_o=WORD，地址4Byte对齐。

coverage check点：
直接用例覆盖，不收功能覆盖率""",
        "note": ""
    },
    {
        "tc_id": "TC_014",
        "tc_name": "tbus_dfx_observability_test",
        "tc_description": """配置条件：
1.仿真环境，模块使能；
2.启用波形采集。

输入激励：
1.执行多种命令序列(WR_CSR、RD_CSR、AHB_WR32、AHB_RD32)；
2.采集状态机状态信号state；
3.采集opcode、addr、wdata、rdata、status_code寄存器；
4.采集rx/tx计数、pcs_n_i、pdi_i、pdo_o、pdo_oe_o信号。

期望结果：
1.所有调试可观测点可采集；
2.便于联调、波形定位和故障复现。

coverage check点：
直接用例覆盖，不收功能覆盖率""",
        "note": ""
    },
    {
        "tc_id": "TC_015",
        "tc_name": "tbus_memory_map_test",
        "tc_description": """配置条件：
1.模块使能；
2.配置SoC AHB总线可访问的memory地址范围。

输入激励：
1.通过WR_CSR(8'h10)/RD_CSR(8'h11)访问模块内部CSR寄存器(VERSION/CTRL/STATUS/LAST_ERR)；
2.发送AHB_WR32(8'h20)/AHB_RD32(8'h21)命令访问SoC内部memory-mapped地址；
3.验证模块自身CSR不进入SoC AHB memory map；
4.验证AHB Master访问采用32bit地址空间和word访问粒度。

期望结果：
1.CSR通过协议命令访问正常；
2.AHB Master访问正确，haddr_o为32bit地址；
3.模块CSR不出现在AHB memory map中。

coverage check点：
直接用例覆盖，不收功能覆盖率""",
        "note": ""
    },
    {
        "tc_id": "TC_016",
        "tc_name": "tbus_ahb_interface_test",
        "tc_description": """配置条件：
1.模块使能，AHB总线空闲；
2.配置hready_i=1，hresp_i=0。

输入激励：
1.发送AHB_WR32(8'h20)命令，检查haddr_o/hwrite_o/htrans_o/hsize_o/hburst_o/hwdata_o输出；
2.发送AHB_RD32(8'h21)命令，检查hrdata_i/hready_i/hresp_i输入处理；
3.模拟hready_i=0等待状态，验证模块正确等待；
4.模拟hresp_i=1错误响应，验证返回STS_AHB_ERR(0x07)；
5.验证htrans_o仅输出IDLE(2'b00)和NONSEQ(2'b10)；
6.验证hsize_o固定为WORD(3'b010)，hburst_o固定为SINGLE(3'b000)。

期望结果：
1.AHB-Lite Master接口符合AMBA规范；
2.单次读写事务正确；
3.信号时序满足AHB协议要求；
4.错误响应正确处理。

coverage check点：
直接用例覆盖，不收功能覆盖率""",
        "note": ""
    }
]

# TP to Checker/Testcase mapping
TP_MAPPING = {
    "TP_001": {"checker": "CHK_001", "testcase": "TC_001"},
    "TP_002": {"checker": "CHK_002", "testcase": "TC_002"},
    "TP_003": {"checker": "CHK_003", "testcase": "TC_003"},
    "TP_004": {"checker": "CHK_004", "testcase": "TC_004"},
    "TP_005": {"checker": "CHK_005", "testcase": "TC_005"},
    "TP_006": {"checker": "CHK_006", "testcase": "TC_006"},
    "TP_007": {"checker": "CHK_007", "testcase": "TC_007"},
    "TP_008": {"checker": "CHK_008", "testcase": "TC_008"},
    "TP_009": {"checker": "CHK_009", "testcase": "TC_009"},
    "TP_010": {"checker": "CHK_010", "testcase": "TC_010"},
    "TP_011": {"checker": "CHK_011", "testcase": "TC_011"},
    "TP_012": {"checker": "CHK_012", "testcase": "TC_012"},
    "TP_013": {"checker": "CHK_013", "testcase": "TC_013"},
    "TP_014": {"checker": "CHK_014", "testcase": "TC_014"},
    "TP_015": {"checker": "CHK_015", "testcase": "TC_015"},
    "TP_016": {"checker": "CHK_016", "testcase": "TC_016"},
}

def copy_cell_style(src_cell, dst_cell):
    """Copy style from source cell to destination cell."""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = copy(src_cell.number_format)
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)

def apply_data_cell_style(cell, col_idx, font_size, is_description=False):
    """Apply standard data cell formatting."""
    from openpyxl.styles import Font, Alignment, Border, Side

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    cell.font = Font(name='宋体', size=font_size)
    cell.border = thin_border

    if is_description:
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    else:
        cell.alignment = Alignment(horizontal='center', vertical='center')

def apply_header_style(cell, is_title=False):
    """Apply header cell formatting with proper fill color."""
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.styles.colors import Color

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Use a light blue fill for header rows (similar to template)
    header_fill = PatternFill(start_color='D6DCE5', end_color='D6DCE5', fill_type='solid')

    cell.font = Font(name='宋体', size=11.0, bold=is_title)
    cell.border = thin_border
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')


def recreate_checker_instruction_section(ws, start_row):
    """Recreate the instruction section for Checker List sheet."""
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Row 1: "填写要求" - merged A:D, bold, header fill
    row = start_row
    cell = ws.cell(row=row, column=1, value="填写要求")
    cell.font = Font(name='宋体', size=11.0, bold=True)
    cell.border = thin_border
    cell.fill = PatternFill(start_color='D6DCE5', end_color='D6DCE5', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in range(2, 5):
        c = ws.cell(row=row, column=col)
        c.border = thin_border
        c.fill = PatternFill(start_color='D6DCE5', end_color='D6DCE5', fill_type='solid')
    ws.merge_cells(f'A{row}:D{row}')

    # Row 2: "CHK Name" instruction
    row = start_row + 1
    cell = ws.cell(row=row, column=1, value="CHK Name")
    cell.font = Font(name='宋体', size=11.0)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

    cell = ws.cell(row=row, column=2, value="如果是SVA checker，需填写SVA property名字")
    cell.font = Font(name='宋体', size=8.0)
    cell.border = thin_border
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    for col in range(3, 5):
        c = ws.cell(row=row, column=col)
        c.border = thin_border
    ws.merge_cells(f'B{row}:D{row}')

    # Row 3: "CHK描述" instruction
    row = start_row + 2
    cell = ws.cell(row=row, column=1, value="CHK描述")
    cell.font = Font(name='宋体', size=11.0)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

    cell = ws.cell(row=row, column=2, value="1、需要定性+定量描述，需具体到check的信号、取值 \n2、和DV SPEC中Checker方案描述的区别 — RTM中描述check的内容，DV SPEC中描述checker的实现方案(例如是通过SVA实现还是scoreboard实时数据比对，还是文件对比)")
    cell.font = Font(name='宋体', size=8.0)
    cell.border = thin_border
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    for col in range(3, 5):
        c = ws.cell(row=row, column=col)
        c.border = thin_border
    ws.merge_cells(f'B{row}:D{row}')


def generate_rtm(template_path, output_path):
    """Generate RTM file with checkers and testcases."""
    wb = openpyxl.load_workbook(template_path)

    # === Update Checker List sheet ===
    ws_checker = wb['Checker List']

    # Template has data rows 3-13 (11 rows), instruction starts at row 14
    # We need 16 checkers, so need to insert 5 rows before row 14
    num_checkers = len(CHECKERS)
    template_data_rows = 11  # rows 3-13
    rows_to_insert = num_checkers - template_data_rows

    # Save instruction section content before modification
    instruction_start_row = 14

    # Clear existing data rows and remove old merged cells in data area
    # Remove merged cells that are in the instruction section area
    merged_to_remove = []
    for mc in ws_checker.merged_cells.ranges:
        if mc.min_row >= instruction_start_row:
            merged_to_remove.append(str(mc))
    for mc_str in merged_to_remove:
        ws_checker.unmerge_cells(mc_str)

    # Clear data rows
    for row in range(3, instruction_start_row):
        for col in range(1, 5):
            ws_checker.cell(row=row, column=col).value = None

    if rows_to_insert > 0:
        # Insert rows before the instruction section
        ws_checker.insert_rows(instruction_start_row, rows_to_insert)

    # Write checkers starting from row 3
    for i, checker in enumerate(CHECKERS):
        row = 3 + i
        # CHK编号 - center, 11pt
        cell = ws_checker.cell(row=row, column=1, value=checker["chk_id"])
        apply_data_cell_style(cell, 1, 11)

        # CHK Name - center, 11pt
        cell = ws_checker.cell(row=row, column=2, value=checker["chk_name"])
        apply_data_cell_style(cell, 2, 11)

        # CHK描述 - left, 8pt, wrap
        cell = ws_checker.cell(row=row, column=3, value=checker["chk_description"])
        apply_data_cell_style(cell, 3, 8, is_description=True)

        # 备注 - center, 11pt
        cell = ws_checker.cell(row=row, column=4, value=checker["note"])
        apply_data_cell_style(cell, 4, 11)

    # Recreate instruction section at the new position
    new_instruction_start = 3 + num_checkers
    recreate_checker_instruction_section(ws_checker, new_instruction_start)

    # === Update DV Testcase List sheet ===
    ws_tc = wb['DV Testcase List']

    # Template has data rows 3-14 (12 rows)
    # We need 16 testcases
    num_testcases = len(TESTCASES)
    template_tc_rows = 12  # rows 3-14

    # Clear existing template testcases
    for row in range(3, 15):
        for col in range(1, 5):
            ws_tc.cell(row=row, column=col).value = None

    # Add rows if needed
    tc_rows_to_insert = num_testcases - template_tc_rows
    if tc_rows_to_insert > 0:
        ws_tc.insert_rows(15, tc_rows_to_insert)

    # Write testcases starting from row 3
    for i, tc in enumerate(TESTCASES):
        row = 3 + i
        # TC编号 - center, 8pt
        cell = ws_tc.cell(row=row, column=1, value=tc["tc_id"])
        apply_data_cell_style(cell, 1, 8)

        # TC Name - center, 8pt
        cell = ws_tc.cell(row=row, column=2, value=tc["tc_name"])
        apply_data_cell_style(cell, 2, 8)

        # TC描述 - left, 6pt, wrap
        cell = ws_tc.cell(row=row, column=3, value=tc["tc_description"])
        apply_data_cell_style(cell, 3, 6, is_description=True)

        # 备注 - center, 11pt
        cell = ws_tc.cell(row=row, column=4, value=tc["note"])
        apply_data_cell_style(cell, 4, 11)

    # === Update FL-TP sheet with checker and testcase links ===
    ws_fl_tp = wb['FL-TP']

    # Update checker_id and testcase_id columns (columns 5 and 6)
    for row in range(3, ws_fl_tp.max_row + 1):
        tp_id = ws_fl_tp.cell(row=row, column=3).value
        if tp_id and tp_id in TP_MAPPING:
            mapping = TP_MAPPING[tp_id]
            cell = ws_fl_tp.cell(row=row, column=5, value=mapping["checker"])
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell = ws_fl_tp.cell(row=row, column=6, value=mapping["testcase"])
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Save the workbook
    wb.save(output_path)
    print(f"RTM file generated: {output_path}")
    print(f"  - Checker List: {num_checkers} checkers written")
    print(f"  - DV Testcase List: {num_testcases} testcases written")
    print(f"  - FL-TP: {len(TP_MAPPING)} TP links updated")
    return output_path

if __name__ == '__main__':
    template_path = 'RTM_AI探索_R3.xlsx'
    output_path = 'RTM_TBUS_Generated.xlsx'

    if not os.path.exists(template_path):
        print(f"Error: Template file not found: {template_path}")
        exit(1)

    generate_rtm(template_path, output_path)
