#!/usr/bin/env python3
"""
Generate Checker and Testcase for RTM based on LRS document
"""

import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from copy import copy

# Checker definitions based on LRS document
CHECKERS = {
    "CHK_001": {
        "name": "clk_freq_checker",
        "description": """1.检查频率值是否正确：连续采集clk_i时钟上升沿和下降沿并计算实际频率/周期，判断是否处于目标频率100MHz的允许误差范围内（±1%，即99MHz~101MHz）；
2.检查时钟稳定性：通过连续采样周期，对比相邻两个周期的频率偏差是否在3%内；
3.检查协议采样、状态机控制和AHB-Lite主接口是否在同域工作，无跨时钟域问题。"""
    },
    "CHK_002": {
        "name": "reset_state_checker",
        "description": """检查rst_ni复位后的模块状态：
1.检查状态机是否回到IDLE状态；
2.检查CTRL.EN=0（模块未使能）；
3.检查CTRL.LANE_MODE=00（1-bit模式）；
4.检查STATUS寄存器值为0x00（无busy、无错误）；
5.检查LAST_ERR寄存器值为0x00；
6.检查协议上下文和响应上下文已清空。"""
    },
    "CHK_003": {
        "name": "csr_access_checker",
        "description": """检查CSR寄存器访问正确性：
1.检查VERSION寄存器（地址0x00）读取返回预期版本值；
2.检查CTRL寄存器（地址0x04）读写正确：EN[0]、LANE_MODE[2:1]、SOFT_RST[3]字段可配置；
3.检查STATUS寄存器（地址0x08）只读属性：BUSY[0]、RESP_VALID[1]、CMD_ERR[2]、BUS_ERR[3]、FRAME_ERR[4]、IN_TEST_MODE[5]、OUT_EN[6]位反映正确状态；
4.检查LAST_ERR寄存器（地址0x0C）只读属性：记录最近失败错误码；
5.检查非法CSR地址访问返回STS_BAD_REG(0x02)。"""
    },
    "CHK_004": {
        "name": "test_mode_checker",
        "description": """检查测试模式工作正确性：
1.检查test_mode_i=1且CTRL.EN=1时模块允许执行功能命令；
2.检查test_mode_i=0时命令返回STS_NOT_IN_TEST(0x05)；
3.检查CTRL.EN=0时命令返回STS_DISABLED(0x04)；
4.检查半双工请求/响应模式：请求阶段pdi_i输入，响应阶段pdo_o输出；
5.检查turnaround周期pdo_oe_o=0（高阻态）。"""
    },
    "CHK_005": {
        "name": "lane_mode_checker",
        "description": """检查lane模式配置正确性：
1.检查LANE_MODE=00时，仅pdi_i[0]/pdo_o[0]有效，其他位忽略；
2.检查LANE_MODE=01时，pdi_i[3:0]/pdo_o[3:0]按高nibble优先发送；
3.检查LANE_MODE=10/11为非法值，返回错误或忽略；
4.检查事务执行期间lane mode不可动态切换；
5.检查SOFT_RST=1后LANE_MODE恢复默认值00（1-bit）。"""
    },
    "CHK_006": {
        "name": "data_interface_checker",
        "description": """检查数据接口时序和格式：
1.检查pcs_n_i帧边界：pcs_n_i=1表示空闲，pcs_n_i=0表示事务开始；
2.检查MSB first传输顺序：高位先发；
3.检查1-bit模式：pdi_i[0]/pdo_o[0]逐位传输；
4.检查4-bit模式：pdi_i[3:0]/pdo_o[3:0]按nibble并行传输；
5.检查pdo_oe_o输出使能：请求阶段为0，响应阶段为1。"""
    },
    "CHK_007": {
        "name": "protocol_timing_checker",
        "description": """检查协议时序正确性：
1.检查turnaround固定1个clk_i周期；
2.检查请求阶段与响应阶段之间pdo_oe_o正确切换；
3.检查四种命令帧格式正确解析：
   - WR_CSR(0x10)：8bit opcode + 8bit addr + 32bit wdata
   - RD_CSR(0x11)：8bit opcode + 8bit addr
   - AHB_WR32(0x20)：8bit opcode + 32bit addr + 32bit wdata
   - AHB_RD32(0x21)：8bit opcode + 32bit addr
4.检查帧长度与opcode匹配，不匹配返回STS_FRAME_ERR(0x06)。"""
    },
    "CHK_008": {
        "name": "opcode_decode_checker",
        "description": """检查opcode译码正确性：
1.检查opcode=0x10正确译码为WR_CSR命令；
2.检查opcode=0x11正确译码为RD_CSR命令；
3.检查opcode=0x20正确译码为AHB_WR32命令；
4.检查opcode=0x21正确译码为AHB_RD32命令；
5.检查非法opcode(非0x10/0x11/0x20/0x21)返回STS_BAD_OPCODE(0x01)；
6.检查前端收满8bit opcode后正确确定期望帧长。"""
    },
    "CHK_009": {
        "name": "ctrl_config_checker",
        "description": """检查CTRL寄存器配置正确性：
1.检查CTRL.EN写1后模块使能，可执行功能命令；
2.检查CTRL.EN写0后模块禁用，命令返回STS_DISABLED(0x04)；
3.检查CTRL.LANE_MODE=00/01配置生效；
4.检查CTRL.SOFT_RST写1后触发软复位：协议上下文清零、LANE_MODE恢复00；
5.检查软复位后STATUS.CMD_ERR、STATUS.BUS_ERR、STATUS.FRAME_ERR清零。"""
    },
    "CHK_010": {
        "name": "status_polling_checker",
        "description": """检查状态查询机制：
1.检查STATUS.BUSY=1表示当前有事务执行；
2.检查STATUS.RESP_VALID=1表示最近响应有效；
3.检查STATUS.CMD_ERR=1表示命令错误（sticky位）；
4.检查STATUS.BUS_ERR=1表示总线错误（sticky位）；
5.检查STATUS.FRAME_ERR=1表示帧错误（sticky位）；
6.检查LAST_ERR记录最近失败错误码（0x01~0x08）；
7.检查无独立中断输出，状态仅通过轮询获取。"""
    },
    "CHK_011": {
        "name": "error_handling_checker",
        "description": """检查异常处理正确性：
1.检查BAD_OPCODE(0x01)：opcode非0x10/0x11/0x20/0x21时返回；
2.检查BAD_REG(0x02)：CSR地址非0x00/0x04/0x08/0x0C时返回；
3.检查ALIGN_ERR(0x03)：AHB地址非4Byte对齐时返回；
4.检查DISABLED(0x04)：CTRL.EN=0时返回；
5.检查NOT_IN_TEST(0x05)：test_mode_i=0时返回；
6.检查FRAME_ERR(0x06)：pcs_n_i提前拉高或帧长度不匹配时返回；
7.检查AHB_ERR(0x07)：hresp_i=1时返回；
8.检查TIMEOUT(0x08)：AHB等待超过BUS_TIMEOUT_CYCLES(默认1024)时返回；
9.检查前置错误在发起AHB访问前完成收敛，不得发起无效总线访问。"""
    },
    "CHK_012": {
        "name": "low_power_checker",
        "description": """检查低功耗设计：
1.检查空闲态（pcs_n_i=1）接收移位寄存器不更新；
2.检查非发送阶段发送移位寄存器不翻转；
3.检查timeout counter仅在AHB等待态工作；
4.检查test_mode_i=0时AHB输出保持IDLE（htrans_o=0）；
5.检查仅RX/TX/WAIT_AHB状态关键寄存器翻转。"""
    },
    "CHK_013": {
        "name": "performance_checker",
        "description": """检查性能指标：
1.检查clk_i=100MHz稳定工作；
2.检查端到端延时：4-bit模式AHB_RD32约23~24 cycles，AHB_WR32约23 cycles；
3.检查吞吐率：1-bit模式约12.5MB/s，4-bit模式约50MB/s；
4.检查AHB接口32bit字访问约束：hsize_o固定为WORD(010)；
5.检查AHB地址4Byte对齐：addr[1:0]=00。"""
    },
    "CHK_014": {
        "name": "dfx_observability_checker",
        "description": """检查DFX可观测性：
1.检查状态机状态可观测；
2.检查opcode、addr、wdata、rdata、status_code可采集；
3.检查rx/tx计数可观测；
4.检查pcs_n_i、pdi_i、pdo_o、pdo_oe_o信号可采集；
5.检查htrans_o、hready_i、hresp_i可观测。"""
    },
    "CHK_015": {
        "name": "memory_map_checker",
        "description": """检查memory map正确性：
1.检查模块CSR（VERSION/CTRL/STATUS/LAST_ERR）仅通过协议命令访问，不进入SoC AHB memory map；
2.检查AHB Master访问使用32bit地址空间；
3.检查AHB Master访问粒度为word（32bit）；
4.检查AHB地址范围由SoC顶层约束。"""
    },
    "CHK_016": {
        "name": "ahb_interface_checker",
        "description": """检查AHB-Lite Master接口符合规范：
1.检查haddr_o[31:0]输出正确的32bit地址；
2.检查hwrite_o：写命令为1，读命令为0；
3.检查htrans_o仅输出IDLE(00)或NONSEQ(10)；
4.检查hsize_o固定为WORD(010)，表示32bit传输；
5.检查hburst_o固定为SINGLE(000)，表示单次传输；
6.检查hwdata_o[31:0]在写数据相输出正确数据；
7.检查hrdata_i[31:0]正确接收读返回数据；
8.检查hready_i=1时数据有效，hready_i=0时等待；
9.检查hresp_i=1时返回STS_AHB_ERR(0x07)。"""
    }
}

# Testcase definitions based on LRS document
TESTCASES = {
    "DV_TC_001": {
        "name": "aplc_clk_freq_test",
        "description": """配置条件：
1.配置clk_i时钟频率为100MHz；
2.配置test_mode_i=1，通过WR_CSR(0x10)写CTRL.EN=1使能模块；
3.配置LANE_MODE=00（1-bit模式）和01（4-bit模式）。

输入激励：
1.连续发送AHB_RD32(0x21)命令，读取SoC内部有效地址；
2.连续发送AHB_WR32(0x20)命令，写入SoC内部有效地址；
3.覆盖1-bit和4-bit两种lane模式。

期望结果：
1.所有命令在100MHz频率下正确执行；
2.协议采样、状态机控制和AHB-Lite主接口同域工作正常；
3.无时序违例，状态码返回0x00（成功）。

coverage check点：
直接用例覆盖，不收功能覆盖率。"""
    },
    "DV_TC_002": {
        "name": "aplc_reset_test",
        "description": """配置条件：
1.初始状态rst_ni为低，模块处于复位状态；
2.配置clk_i=100MHz。

输入激励：
1.释放rst_ni，验证复位后状态机回到IDLE；
2.通过RD_CSR(0x11)读取CTRL寄存器，验证EN=0、LANE_MODE=00；
3.通过RD_CSR(0x11)读取STATUS寄存器，验证值为0x00；
4.通过RD_CSR(0x11)读取LAST_ERR寄存器，验证值为0x00。

期望结果：
1.复位后状态机处于IDLE状态；
2.CTRL.EN=0（模块未使能）；
3.CTRL.LANE_MODE=00（1-bit模式）；
4.STATUS=0x00，LAST_ERR=0x00，无残留状态。

coverage check点：
直接用例覆盖，不收功能覆盖率。"""
    },
    "DV_TC_003": {
        "name": "aplc_csr_access_test",
        "description": """配置条件：
1.配置test_mode_i=1，通过WR_CSR(0x10)写CTRL.EN=1使能模块；
2.配置LANE_MODE=01（4-bit模式）。

输入激励：
1.发送RD_CSR(0x11)命令读取VERSION寄存器（地址0x00），验证版本号；
2.发送WR_CSR(0x10)命令写CTRL寄存器（地址0x04），配置EN=1、LANE_MODE=01、SOFT_RST=0；
3.发送RD_CSR(0x11)命令读取CTRL寄存器，验证写入正确；
4.发送RD_CSR(0x11)命令读取STATUS寄存器（地址0x08），检查BUSY、RESP_VALID等位；
5.发送RD_CSR(0x11)命令读取LAST_ERR寄存器（地址0x0C）；
6.发送RD_CSR(0x11)命令访问非法CSR地址（如0x10），验证返回STS_BAD_REG(0x02)。

期望结果：
1.VERSION寄存器返回预期版本值；
2.CTRL寄存器读写正确，配置生效；
3.STATUS寄存器反映正确状态；
4.LAST_ERR寄存器记录错误码；
5.非法CSR地址返回STS_BAD_REG(0x02)。

coverage check点：
直接用例覆盖，不收功能覆盖率。"""
    },
    "DV_TC_004": {
        "name": "aplc_test_mode_test",
        "description": """配置条件：
1.配置clk_i=100MHz，test_mode_i=1；
2.通过WR_CSR(0x10)写CTRL.EN=1使能模块。

输入激励：
1.发送WR_CSR(0x10)命令配置LANE_MODE=01（4-bit模式）；
2.发送AHB_WR32(0x20)命令进行单次写操作；
3.发送AHB_RD32(0x21)命令进行单次读操作；
4.验证半双工请求/响应模式工作正常；
5.验证turnaround周期pdo_oe_o=0。

期望结果：
1.模块正确执行WR_CSR、AHB_WR32、AHB_RD32功能命令；
2.协议采用半双工模式：请求阶段pdi_i输入，响应阶段pdo_o输出；
3.turnaround固定1周期，pdo_oe_o正确切换；
4.状态码返回0x00（成功）。

coverage check点：
直接用例覆盖，不收功能覆盖率。"""
    },
    "DV_TC_005": {
        "name": "aplc_lane_mode_test",
        "description": """配置条件：
1.配置test_mode_i=1，CTRL.EN=1使能模块；
2.依次配置不同LANE_MODE。

输入激励：
1.配置CTRL.LANE_MODE=00（1-bit模式），发送AHB_RD32命令，验证pdi_i[0]/pdo_o[0]逐位传输；
2.配置CTRL.LANE_MODE=01（4-bit模式），发送AHB_RD32命令，验证pdi_i[3:0]/pdo_o[3:0]并行传输，高nibble优先；
3.尝试配置LANE_MODE=10/11，验证返回错误或忽略；
4.验证事务执行期间lane mode不可动态切换；
5.配置CTRL.SOFT_RST=1触发软复位，验证LANE_MODE恢复00。

期望结果：
1.1-bit模式使用bit0传输，其他位忽略；
2.4-bit模式使用pdi_i[3:0]/pdo_o[3:0]并行传输；
3.LANE_MODE=10/11为非法值；
4.软复位后LANE_MODE恢复默认值00；
5.Burst和AXI功能确认不可用。

coverage check点：
直接用例覆盖，不收功能覆盖率。"""
    },
    "DV_TC_006": {
        "name": "aplc_data_interface_test",
        "description": """配置条件：
1.配置test_mode_i=1，CTRL.EN=1使能模块；
2.配置LANE_MODE=00（1-bit）和01（4-bit）两种模式。

输入激励：
1.在1-bit模式下发送WR_CSR(0x10)命令，验证pdi_i[0]/pdo_o[0]数据传输；
2.在1-bit模式下发送AHB_WR32(0x20)命令，验证帧边界和数据顺序；
3.在4-bit模式下发送相同命令，验证pdi_i[3:0]/pdo_o[3:0]按高nibble优先发送；
4.验证pcs_n_i帧边界：pcs_n_i=0开启事务，pcs_n_i=1结束事务；
5.验证MSB first传输顺序。

期望结果：
1.1-bit模式仅bit0有效，其他位忽略；
2.4-bit模式按nibble并行传输；
3.pcs_n_i帧边界清晰正确；
4.MSB first顺序正确；
5.pdo_oe_o在响应阶段为1。

coverage check点：
直接用例覆盖，不收功能覆盖率。"""
    },
    "DV_TC_007": {
        "name": "aplc_protocol_timing_test",
        "description": """配置条件：
1.配置test_mode_i=1，CTRL.EN=1使能模块；
2.配置LANE_MODE=01（4-bit模式）。

输入激励：
1.发送WR_CSR(0x10)命令验证请求阶段；
2.验证turnaround周期为固定1个clk_i周期；
3.检查响应阶段pdo_oe_o=1输出使能；
4.验证四种命令帧格式：
   - WR_CSR(0x10)：8bit opcode + 8bit addr + 32bit wdata = 48bit = 12拍（4-bit模式）
   - RD_CSR(0x11)：8bit opcode + 8bit addr = 16bit = 4拍（4-bit模式）
   - AHB_WR32(0x20)：8bit opcode + 32bit addr + 32bit wdata = 72bit = 18拍（4-bit模式）
   - AHB_RD32(0x21)：8bit opcode + 32bit addr = 40bit = 10拍（4-bit模式）；
5.提前拉高pcs_n_i验证FRAME_ERR检测。

期望结果：
1.请求/响应阶段正确分离；
2.turnaround固定1周期；
3.协议字段MSB first传输；
4.四种帧格式正确处理；
5.帧长度不匹配返回STS_FRAME_ERR(0x06)。

coverage check点：
直接用例覆盖，不收功能覆盖率。"""
    },
    "DV_TC_008": {
        "name": "aplc_opcode_test",
        "description": """配置条件：
1.配置test_mode_i=1，CTRL.EN=1使能模块；
2.配置AHB总线空闲正常。

输入激励：
1.发送opcode=0x10执行WR_CSR命令，验证正确译码；
2.发送opcode=0x11执行RD_CSR命令，验证正确译码；
3.发送opcode=0x20执行AHB_WR32命令，验证正确译码；
4.发送opcode=0x21执行AHB_RD32命令，验证正确译码；
5.发送非法opcode（如0x00、0xFF、0x30等非0x10/0x11/0x20/0x21值）；
6.验证前端收满8bit opcode后正确确定期望帧长。

期望结果：
1.四种opcode正确译码并执行对应功能；
2.非法opcode返回STS_BAD_OPCODE(0x01)；
3.LAST_ERR更新为0x01；
4.STATUS.CMD_ERR=1；
5.前端正确确定期望帧长。

coverage check点：
直接用例覆盖，不收功能覆盖率。"""
    },
    "DV_TC_009": {
        "name": "aplc_ctrl_config_test",
        "description": """配置条件：
1.配置test_mode_i=1；
2.初始CTRL.EN=0。

输入激励：
1.写CTRL.EN=1使能模块，发送AHB_RD32命令验证功能命令可执行；
2.写CTRL.EN=0禁用模块，发送命令验证返回STS_DISABLED(0x04)；
3.写CTRL.LANE_MODE=00验证1-bit模式生效；
4.写CTRL.LANE_MODE=01验证4-bit模式生效；
5.写CTRL.SOFT_RST=1触发软复位，验证协议上下文清零、LANE_MODE恢复00；
6.验证软复位后STATUS寄存器错误位清零。

期望结果：
1.CTRL.EN使能/禁用状态切换正确；
2.CTRL.LANE_MODE配置生效；
3.软复位功能正常；
4.协议上下文清零；
5.LANE_MODE恢复默认值00。

coverage check点：
直接用例覆盖，不收功能覆盖率。"""
    },
    "DV_TC_010": {
        "name": "aplc_status_polling_test",
        "description": """配置条件：
1.配置test_mode_i=1，CTRL.EN=1使能模块；
2.配置各类错误触发条件。

输入激励：
1.触发非法opcode错误，通过RD_CSR读取STATUS寄存器，验证CMD_ERR=1；
2.触发AHB错误（hresp_i=1），通过RD_CSR读取STATUS寄存器，验证BUS_ERR=1；
3.触发帧错误，通过RD_CSR读取STATUS寄存器，验证FRAME_ERR=1；
4.执行命令期间读取STATUS.BUSY=1；
5.命令完成后读取STATUS.RESP_VALID=1；
6.读取LAST_ERR寄存器获取最近错误码；
7.验证无独立中断输出引脚。

期望结果：
1.STATUS寄存器各sticky位正确置位；
2.LAST_ERR记录最近失败错误码；
3.MVP版本无中断输出；
4.ATE轮询方式正确获取状态。

coverage check点：
直接用例覆盖，不收功能覆盖率。"""
    },
    "DV_TC_011": {
        "name": "aplc_error_handling_test",
        "description": """配置条件：
1.配置test_mode_i=1，CTRL.EN=1使能模块；
2.配置AHB总线正常响应。

输入激励：
1.发送非法opcode（0x00/0xFF）验证返回STS_BAD_OPCODE(0x01)；
2.访问非法CSR地址（0x10等）验证返回STS_BAD_REG(0x02)；
3.发送非4Byte对齐地址（如0x00000001）执行AHB命令，验证返回STS_ALIGN_ERR(0x03)；
4.配置CTRL.EN=0，发送命令验证返回STS_DISABLED(0x04)；
5.配置test_mode_i=0，发送命令验证返回STS_NOT_IN_TEST(0x05)；
6.发送命令期间提前拉高pcs_n_i，验证返回STS_FRAME_ERR(0x06)；
7.触发hresp_i=1，验证返回STS_AHB_ERR(0x07)；
8.模拟AHB超时（hready_i=0持续超过1024周期），验证返回STS_TIMEOUT(0x08)；
9.验证前置错误在发起AHB访问前完成收敛。

期望结果：
1.各类错误正确检测并返回对应状态码；
2.LAST_ERR更新正确；
3.STATUS对应错误位置位；
4.不发起无效总线访问。

coverage check点：
直接用例覆盖，不收功能覆盖率。"""
    },
    "DV_TC_012": {
        "name": "aplc_low_power_test",
        "description": """配置条件：
1.模块处于空闲态（pcs_n_i=1或test_mode_i=0或无有效任务）；
2.配置clk_i=100MHz。

输入激励：
1.保持pcs_n_i=1，监控接收移位寄存器是否翻转；
2.保持非发送状态，监控发送移位寄存器是否翻转；
3.模拟AHB等待态（hready_i=0），验证timeout counter工作；
4.配置test_mode_i=0，验证AHB输出保持IDLE（htrans_o=0）；
5.切换到工作态，监控关键寄存器翻转情况。

期望结果：
1.空闲态接收/发送移位寄存器不翻转；
2.timeout counter仅在AHB等待态工作；
3.test_mode_i=0时AHB输出保持IDLE；
4.仅RX/TX/WAIT_AHB状态关键寄存器翻转。

coverage check点：
直接用例覆盖，不收功能覆盖率。"""
    },
    "DV_TC_013": {
        "name": "aplc_performance_test",
        "description": """配置条件：
1.配置clk_i=100MHz；
2.配置test_mode_i=1，CTRL.EN=1使能模块；
3.配置AHB总线响应正常（hready_i=1）。

输入激励：
1.在1-bit模式下连续执行AHB_RD32/AHB_WR32命令，计算端到端延时和吞吐率；
2.在4-bit模式下执行相同测试；
3.验证链路线速：1-bit模式约12.5MB/s，4-bit模式约50MB/s；
4.验证AHB接口32bit字访问（hsize_o=010）；
5.验证AHB地址4Byte对齐（addr[1:0]=00）。

期望结果：
1.目标频率100MHz稳定工作；
2.端到端延时：4-bit模式AHB_RD32约23~24 cycles，AHB_WR32约23 cycles；
3.吞吐率达标；
4.AHB访问约束满足。

coverage check点：
直接用例覆盖，不收功能覆盖率。"""
    },
    "DV_TC_014": {
        "name": "aplc_dfx_test",
        "description": """配置条件：
1.仿真环境，test_mode_i=1，CTRL.EN=1使能模块。

输入激励：
1.执行多种命令序列（WR_CSR、RD_CSR、AHB_WR32、AHB_RD32）；
2.采集状态机状态、opcode、addr、wdata、rdata、status_code；
3.采集rx/tx计数；
4.采集pcs_n_i、pdi_i、pdo_o、pdo_oe_o信号；
5.采集htrans_o、hready_i、hresp_i信号。

期望结果：
1.所有调试可观测点可采集；
2.便于联调、波形定位和故障复现；
3.流量统计可观测。

coverage check点：
直接用例覆盖，不收功能覆盖率。"""
    },
    "DV_TC_015": {
        "name": "aplc_memory_map_test",
        "description": """配置条件：
1.配置test_mode_i=1，CTRL.EN=1使能模块。

输入激励：
1.通过WR_CSR/RD_CSR访问模块内部CSR寄存器（VERSION地址0x00、CTRL地址0x04、STATUS地址0x08、LAST_ERR地址0x0C）；
2.发送AHB_WR32/AHB_RD32命令访问SoC内部memory-mapped地址；
3.验证模块自身CSR不进入SoC AHB memory map；
4.验证AHB Master访问使用32bit地址空间；
5.验证AHB Master访问粒度为word（32bit）。

期望结果：
1.CSR通过协议命令访问正常；
2.AHB Master访问正确；
3.CSR不暴露在AHB memory map中；
4.地址空间约束满足。

coverage check点：
直接用例覆盖，不收功能覆盖率。"""
    },
    "DV_TC_016": {
        "name": "aplc_ahb_interface_test",
        "description": """配置条件：
1.配置test_mode_i=1，CTRL.EN=1使能模块；
2.配置AHB总线空闲正常。

输入激励：
1.发送AHB_WR32(0x20)命令，验证haddr_o[31:0]输出正确的32bit地址；
2.验证hwrite_o=1（写操作）；
3.验证htrans_o=NONSEQ(10)；
4.验证hsize_o=WORD(010)表示32bit传输；
5.验证hburst_o=SINGLE(000)表示单次传输；
6.验证hwdata_o[31:0]输出正确写数据；
7.发送AHB_RD32(0x21)命令，验证hwrite_o=0（读操作）；
8.验证hrdata_i[31:0]正确接收读返回数据；
9.模拟hready_i=0等待，验证模块正确等待；
10.模拟hresp_i=1，验证返回STS_AHB_ERR(0x07)。

期望结果：
1.AHB-Lite Master接口符合规范；
2.单次读写事务正确；
3.htrans_o仅输出IDLE/NONSEQ；
4.hsize_o固定为WORD；
5.hburst_o固定为SINGLE；
6.错误响应正确处理。

coverage check点：
直接用例覆盖，不收功能覆盖率。"""
    }
}


def generate_rtm(input_file, output_file):
    """Generate new RTM file with Checker and Testcase filled"""

    # Load the input workbook
    wb = openpyxl.load_workbook(input_file)

    # Update Checker List sheet
    checker_sheet = wb['Checker List']

    # Unmerge all cells in Checker List sheet to allow editing
    merged_ranges = list(checker_sheet.merged_cells.ranges)
    for merged_range in merged_ranges:
        checker_sheet.unmerge_cells(str(merged_range))

    # Find the "填写要求" row to know where to stop
    fill_require_row = None
    for row_idx in range(1, checker_sheet.max_row + 1):
        cell_value = checker_sheet.cell(row=row_idx, column=1).value
        if cell_value == "填写要求":
            fill_require_row = row_idx
            break

    # Clear existing data rows (rows 3 to before 填写要求)
    if fill_require_row:
        for row_idx in range(3, fill_require_row):
            for col in range(1, 5):
                checker_sheet.cell(row=row_idx, column=col).value = None

    # Add new checkers starting from row 3 (after header)
    row_idx = 3
    for chk_id in sorted(CHECKERS.keys()):
        checker = CHECKERS[chk_id]
        checker_sheet.cell(row=row_idx, column=1, value=chk_id)
        checker_sheet.cell(row=row_idx, column=2, value=checker['name'])
        checker_sheet.cell(row=row_idx, column=3, value=checker['description'])
        checker_sheet.cell(row=row_idx, column=4, value='')  # 备注

        # Set alignment
        for col in range(1, 5):
            target_cell = checker_sheet.cell(row=row_idx, column=col)
            target_cell.alignment = Alignment(wrap_text=True, vertical='top')

        row_idx += 1

    # Re-merge cells if needed
    # Merge the header cell
    checker_sheet.merge_cells('A1:D1')

    # Update DV Testcase List sheet
    tc_sheet = wb['DV Testcase List']

    # Unmerge all cells in DV Testcase List sheet
    merged_ranges = list(tc_sheet.merged_cells.ranges)
    for merged_range in merged_ranges:
        tc_sheet.unmerge_cells(str(merged_range))

    # Find the last row for testcases
    # Clear existing data rows starting from row 3
    for row_idx in range(3, tc_sheet.max_row + 1):
        cell_value = tc_sheet.cell(row=row_idx, column=1).value
        if cell_value and "填写" in str(cell_value):
            break
        for col in range(1, 5):
            tc_sheet.cell(row=row_idx, column=col).value = None

    # Add new testcases starting from row 3 (after header)
    row_idx = 3
    for tc_id in sorted(TESTCASES.keys()):
        testcase = TESTCASES[tc_id]
        tc_sheet.cell(row=row_idx, column=1, value=tc_id)
        tc_sheet.cell(row=row_idx, column=2, value=testcase['name'])
        tc_sheet.cell(row=row_idx, column=3, value=testcase['description'])
        tc_sheet.cell(row=row_idx, column=4, value='')  # 备注

        # Set alignment
        for col in range(1, 5):
            target_cell = tc_sheet.cell(row=row_idx, column=col)
            target_cell.alignment = Alignment(wrap_text=True, vertical='top')

        row_idx += 1

    # Re-merge header cell
    tc_sheet.merge_cells('A1:D1')

    # Save the workbook
    wb.save(output_file)
    print(f"Generated RTM file: {output_file}")
    print(f"Added {len(CHECKERS)} checkers and {len(TESTCASES)} testcases")


if __name__ == '__main__':
    import sys
    if len(sys.argv) >= 3:
        input_file = sys.argv[1]
        output_file = sys.argv[2]
    else:
        input_file = 'RTM_AI探索_R3(1).xlsx'
        output_file = 'RTM_AI_generated.xlsx'

    generate_rtm(input_file, output_file)
