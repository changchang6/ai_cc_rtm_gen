#!/usr/bin/env python3
"""
RTM Generator Script
Generate Checker List and DV Testcase List for RTM file.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy
import sys
import os

# Key design information extracted from LRS
INTERFACE_SIGNALS = {
    'pcs_n_i': {'dir': 'In', 'width': 1, 'desc': '片选，低有效'},
    'pdi_i': {'dir': 'In', 'width': 4, 'desc': '输入数据'},
    'pdo_o': {'dir': 'Out', 'width': 4, 'desc': '输出数据'},
    'pdo_oe_o': {'dir': 'Out', 'width': 1, 'desc': '输出使能'},
    'clk_i': {'dir': 'In', 'width': 1, 'desc': '主时钟'},
    'rst_ni': {'dir': 'In', 'width': 1, 'desc': '低有效复位'},
    'test_mode_i': {'dir': 'In', 'width': 1, 'desc': '测试模式使能'},
    'haddr_o': {'dir': 'Out', 'width': 32, 'desc': 'AHB地址'},
    'hwrite_o': {'dir': 'Out', 'width': 1, 'desc': 'AHB读写标志'},
    'htrans_o': {'dir': 'Out', 'width': 2, 'desc': 'AHB传输类型'},
    'hsize_o': {'dir': 'Out', 'width': 3, 'desc': 'AHB传输大小'},
    'hburst_o': {'dir': 'Out', 'width': 3, 'desc': 'AHB突发类型'},
    'hwdata_o': {'dir': 'Out', 'width': 32, 'desc': 'AHB写数据'},
    'hrdata_i': {'dir': 'In', 'width': 32, 'desc': 'AHB读数据'},
    'hready_i': {'dir': 'In', 'width': 1, 'desc': 'AHB就绪'},
    'hresp_i': {'dir': 'In', 'width': 1, 'desc': 'AHB响应'},
}

OPCODES = {
    '0x10': {'name': 'WR_CSR', 'desc': 'CSR写'},
    '0x11': {'name': 'RD_CSR', 'desc': 'CSR读'},
    '0x20': {'name': 'AHB_WR32', 'desc': 'AHB单次写'},
    '0x21': {'name': 'AHB_RD32', 'desc': 'AHB单次读'},
}

REGISTERS = {
    'CTRL': {'fields': ['EN', 'LANE_MODE', 'SOFT_RST'], 'desc': '控制寄存器'},
    'STATUS': {'fields': ['BUSY', 'STICKY'], 'desc': '状态寄存器'},
    'LAST_ERR': {'fields': ['ERR_CODE'], 'desc': '最后错误码'},
    'VERSION': {'fields': ['MAJOR', 'MINOR'], 'desc': '版本寄存器'},
}

ERROR_CODES = {
    '0x01': 'STS_BAD_OPCODE',
    '0x02': 'STS_BAD_REG',
    '0x03': 'STS_ALIGN_ERR',
    '0x04': 'STS_DISABLED',
    '0x05': 'STS_NOT_IN_TEST',
    '0x06': 'STS_FRAME_ERR',
    '0x07': 'STS_AHB_ERR',
    '0x08': 'STS_TIMEOUT',
    '0x00': 'STS_OK',
}

# Checker definitions for each TP
CHECKERS = {
    'TP_001': {
        'id': 'CHK_001',
        'name': 'clk_domain_checker',
        'description': '''1. 检查clk_i频率稳定性：在连续时钟周期内，上升沿和下降沿时间间隔应在目标频率100MHz的±3%误差范围内
2. 验证协议采样、状态机控制和AHB-Lite主接口同域工作：检查所有内部寄存器和状态机仅在clk_i上升沿更新
3. 检查无时钟域交叉问题：确认无异步信号直接使用，所有跨域信号均经过同步处理'''
    },
    'TP_002': {
        'id': 'CHK_002',
        'name': 'rst_checker',
        'description': '''1. 检查rst_ni复位连接性：rst_ni拉低后，状态机应立即回到IDLE状态
2. 验证复位后默认值：CTRL.EN=0，CTRL.LANE_MODE=00(1-bit模式)，STATUS=0x00，LAST_ERR=0x00
3. 检查复位同步释放：内部同步释放逻辑确保复位撤销时不产生亚稳态
4. 验证协议上下文清空：rx_shift_reg、tx_shift_reg、opcode、addr、wdata等寄存器复位为0'''
    },
    'TP_003': {
        'id': 'CHK_003',
        'name': 'csr_checker',
        'description': '''1. 检查VERSION寄存器读取：通过opcode=0x11(RD_CSR)读取VERSION，验证返回值符合设计预期
2. 检查CTRL寄存器读写：
   - 写CTRL.EN=1，验证模块使能生效
   - 写CTRL.LANE_MODE=00/01，验证lane模式切换
   - 写CTRL.SOFT_RST=1，验证软复位触发
3. 检查STATUS寄存器：验证BUSY位在任务执行期间置1，完成后清0
4. 检查LAST_ERR寄存器：验证错误发生时LAST_ERR更新为对应错误码'''
    },
    'TP_004': {
        'id': 'CHK_004',
        'name': 'mode_checker',
        'description': '''1. 检查test_mode_i=1且CTRL.EN=1时模块允许执行功能命令
2. 验证test_mode_i=0时发送命令返回错误码STS_NOT_IN_TEST(0x05)
3. 验证CTRL.EN=0时发送命令返回错误码STS_DISABLED(0x04)
4. 检查半双工请求/响应模式：pcs_n_i拉低开始请求阶段，turnaround 1周期后进入响应阶段'''
    },
    'TP_005': {
        'id': 'CHK_005',
        'name': 'lane_mode_checker',
        'description': '''1. 检查CTRL.LANE_MODE=00时仅使用pdi_i[0]/pdo_o[0]传输，每拍收发1bit(MSB first)
2. 验证CTRL.LANE_MODE=01时使用pdi_i[3:0]/pdo_o[3:0]传输，每拍收发4bit(高nibble优先)
3. 确认非法LANE_MODE值(10/11)不改变当前配置或返回错误
4. 检查lane模式切换时序：模式切换应在帧边界进行，不影响当前帧传输'''
    },
    'TP_006': {
        'id': 'CHK_006',
        'name': 'data_interface_checker',
        'description': '''1. 检查pcs_n_i帧边界：pcs_n_i拉低标志帧开始，拉高标志帧结束
2. 验证pdi_i[3:0]输入时序：输入数据在clk_i上升沿采样，采样窗口内数据应稳定
3. 验证pdo_o[3:0]输出时序：输出数据在响应阶段有效，pdo_oe_o=1时输出使能
4. 检查pdo_oe_o方向控制：请求阶段pdo_oe_o=0(输入方向)，响应阶段pdo_oe_o=1(输出方向)
5. 验证MSB first传输顺序：数据从最高位开始传输'''
    },
    'TP_007': {
        'id': 'CHK_007',
        'name': 'protocol_checker',
        'description': '''1. 检查turnaround周期：请求阶段结束后固定1个clk_i周期的turnaround
2. 验证帧格式：
   - WR_CSR(0x10): opcode(8bit) + reg_addr(8bit) + wdata(32bit)
   - RD_CSR(0x11): opcode(8bit) + reg_addr(8bit) -> STATUS(8bit) + rdata(32bit)
   - AHB_WR32(0x20): opcode(8bit) + addr(32bit) + wdata(32bit)
   - AHB_RD32(0x21): opcode(8bit) + addr(32bit) -> STATUS(8bit) + rdata(32bit)
3. 检查协议字段MSB first传输顺序'''
    },
    'TP_008': {
        'id': 'CHK_008',
        'name': 'opcode_checker',
        'description': '''1. 检查opcode解析正确性：
   - opcode=0x10解析为WR_CSR命令
   - opcode=0x11解析为RD_CSR命令
   - opcode=0x20解析为AHB_WR32命令
   - opcode=0x21解析为AHB_RD32命令
2. 验证非法opcode处理：opcode为0x00/0xFF或其他非法值时返回STS_BAD_OPCODE(0x01)
3. 检查帧长确定：前端收满8bit opcode后正确确定期望帧长'''
    },
    'TP_009': {
        'id': 'CHK_009',
        'name': 'ctrl_checker',
        'description': '''1. 检查CTRL.EN配置生效：CTRL.EN=1时模块使能，CTRL.EN=0时模块禁用
2. 验证CTRL.LANE_MODE配置：
   - LANE_MODE=00配置为1-bit模式
   - LANE_MODE=01配置为4-bit模式
   - LANE_MODE=10/11为非法配置
3. 检查CTRL.SOFT_RST功能：写SOFT_RST=1触发协议上下文清零并恢复默认lane模式
4. 验证配置立即生效：配置写入后下一拍生效'''
    },
    'TP_010': {
        'id': 'CHK_010',
        'name': 'status_checker',
        'description': '''1. 检查STATUS寄存器查询：通过RD_CSR读取STATUS，验证BUSY位和STICKY位正确反映模块状态
2. 验证LAST_ERR寄存器查询：错误发生时LAST_ERR更新为对应错误码并保持
3. 确认无独立中断输出：MVP版本不实现中断信号，所有状态通过寄存器查询获取
4. 检查ATE轮询方式：STATUS和LAST_ERR可被ATE主机持续轮询'''
    },
    'TP_011': {
        'id': 'CHK_011',
        'name': 'error_checker',
        'description': '''1. 检查错误码生成正确性：
   - 非法opcode -> STS_BAD_OPCODE(0x01)
   - 非法CSR地址 -> STS_BAD_REG(0x02)
   - 非对齐地址 -> STS_ALIGN_ERR(0x03)
   - CTRL.EN=0 -> STS_DISABLED(0x04)
   - test_mode_i=0 -> STS_NOT_IN_TEST(0x05)
   - pcs_n_i提前拉高 -> STS_FRAME_ERR(0x06)
   - AHB错误(hresp_i=1) -> STS_AHB_ERR(0x07)
   - AHB超时 -> STS_TIMEOUT(0x08)
2. 验证前置错误收敛：在发起AHB访问前完成所有前置错误检测
3. 检查LAST_ERR更新：错误发生时LAST_ERR寄存器更新为对应错误码'''
    },
    'TP_012': {
        'id': 'CHK_012',
        'name': 'lowpower_checker',
        'description': '''1. 检查空闲态寄存器不更新：pcs_n_i=1(空闲)时rx_shift_reg和tx_shift_reg不翻转
2. 验证超时计数器控制：超时计数器仅在WAIT_AHB状态工作时更新
3. 检查AHB输出空闲：test_mode_i=0或无有效任务时AHB输出保持IDLE(htrans_o=IDLE)
4. 验证关键寄存器翻转：仅在RX/TX/WAIT_AHB相关状态翻转关键寄存器'''
    },
    'TP_013': {
        'id': 'CHK_013',
        'name': 'perf_checker',
        'description': '''1. 检查目标频率达标：clk_i在100MHz频率下稳定工作
2. 验证端到端延时：从pcs_n_i拉低到pdo_oe_o拉高的延时约23~24个clk_i周期
3. 检查吞吐率：
   - 1-bit模式原始链路线速：12.5MB/s
   - 4-bit模式原始链路线速：50MB/s
4. 验证AHB访问约束：hsize_o固定为WORD(32-bit)，地址需4Byte对齐'''
    },
    'TP_014': {
        'id': 'CHK_014',
        'name': 'dfx_checker',
        'description': '''1. 检查状态机状态可观测：状态机当前状态可通过调试接口观测
2. 验证关键寄存器可观测：opcode、addr、wdata、rdata、status_code可通过调试接口观测
3. 检查计数器可观测：rx_count、tx_count可通过调试接口观测
4. 验证接口信号可观测：pcs_n_i、pdi_i、pdo_o、pdo_oe_o可通过调试接口观测'''
    },
    'TP_015': {
        'id': 'CHK_015',
        'name': 'memmap_checker',
        'description': '''1. 检查CSR访问方式：模块内部CSR(VERSION/CTRL/STATUS/LAST_ERR)不进入SoC AHB memory map，仅通过协议命令访问
2. 验证AHB Master地址空间：AHB_WR32/AHB_RD32命令访问SoC内部memory-mapped地址，采用32bit地址空间
3. 检查访问粒度：AHB Master访问固定为word(32-bit)访问粒度
4. 验证地址范围约束：具体可达目标地址范围由SoC顶层统一约束'''
    },
    'TP_016': {
        'id': 'CHK_016',
        'name': 'ahb_checker',
        'description': '''1. 检查AHB-Lite Master信号输出：
   - haddr_o[31:0]：目标地址输出
   - hwrite_o：读写标志(1=写，0=读)
   - htrans_o[1:0]：仅输出IDLE(2'b00)和NONSEQ(2'b10)
   - hsize_o[2:0]：固定为WORD(3'b010)
   - hburst_o[2:0]：固定为SINGLE(3'b000)
   - hwdata_o[31:0]：写数据输出
2. 验证AHB输入响应处理：
   - hrdata_i[31:0]：读数据输入
   - hready_i：等待状态处理
   - hresp_i：错误响应(返回STS_AHB_ERR)
3. 检查AHB时序：地址相和数据相符合AHB-Lite协议要求'''
    },
}

# Testcase definitions for each TP
TESTCASES = {
    'TP_001': {
        'id': 'TC_001',
        'name': 'test_clk_domain',
        'description': '''TC场景：测试时钟域同步性
配置条件：
1. test_mode_i=1（测试模式使能）
2. CTRL.EN=1（模块使能）
3. clk_i频率配置为100MHz

输入激励：
1. 连续发送AHB_RD32(opcode=0x21)命令，覆盖1-bit模式(LANE_MODE=00)和4-bit模式(LANE_MODE=01)
2. 连续发送AHB_WR32(opcode=0x20)命令，覆盖两种lane模式
3. 在命令执行过程中监测内部寄存器和状态机更新时机

期望结果：
1. 所有命令在100MHz频率下正确执行
2. 协议采样、状态机控制和AHB-Lite主接口同域工作正常
3. 所有内部寄存器仅在clk_i上升沿更新
4. STATUS=0x00（成功）返回

coverage check点：
对clk_i频率范围(90MHz~110MHz)收集功能覆盖率。'''
    },
    'TP_002': {
        'id': 'TC_002',
        'name': 'test_reset',
        'description': '''TC场景：测试复位功能
配置条件：
1. rst_ni初始为低，模块处于复位状态

输入激励：
1. 释放rst_ni，验证复位释放效果
2. 通过RD_CSR(opcode=0x11)读取CTRL寄存器，检查默认值
3. 通过RD_CSR读取STATUS和LAST_ERR寄存器，检查清零状态
4. 监测状态机状态，确认回到IDLE

期望结果：
1. 复位后状态机处于IDLE状态
2. CTRL.EN=0，CTRL.LANE_MODE=00(1-bit模式)
3. STATUS=0x00，LAST_ERR=0x00
4. 协议上下文(rx_shift_reg, tx_shift_reg等)清零

coverage check点：
直接用例覆盖，不收功能覆盖率。'''
    },
    'TP_003': {
        'id': 'TC_003',
        'name': 'test_csr_access',
        'description': '''TC场景：测试CSR寄存器访问
配置条件：
1. test_mode_i=1
2. CTRL.EN=1
3. LANE_MODE=00(1-bit模式)或01(4-bit模式)

输入激励：
1. 通过RD_CSR(opcode=0x11)读取VERSION寄存器，验证版本号
2. 通过WR_CSR(opcode=0x10)写入CTRL.EN=1，验证模块使能
3. 通过WR_CSR写入CTRL.LANE_MODE=00/01，验证lane模式配置
4. 通过WR_CSR写入CTRL.SOFT_RST=1，验证软复位功能
5. 通过RD_CSR读取STATUS和LAST_ERR，检查状态正确

期望结果：
1. VERSION返回预期版本号
2. CTRL配置正确写入并生效
3. STATUS正确反映BUSY状态
4. LAST_ERR在错误发生时更新

coverage check点：
对CSR寄存器地址范围、LANE_MODE配置值收集功能覆盖率。'''
    },
    'TP_004': {
        'id': 'TC_004',
        'name': 'test_work_mode',
        'description': '''TC场景：测试工作模式
配置条件：
1. 初始状态test_mode_i=0，CTRL.EN=0

输入激励：
1. test_mode_i=0时发送WR_CSR命令，验证返回STS_NOT_IN_TEST(0x05)
2. 设置test_mode_i=1，CTRL.EN=0时发送命令，验证返回STS_DISABLED(0x04)
3. 设置CTRL.EN=1，发送WR_CSR配置lane模式
4. 发送AHB_WR32/AHB_RD32命令验证半双工模式

期望结果：
1. test_mode_i=0时返回STS_NOT_IN_TEST(0x05)
2. CTRL.EN=0时返回STS_DISABLED(0x04)
3. test_mode_i=1且CTRL.EN=1时命令正常执行
4. 半双工请求/响应模式工作正常

coverage check点：
对test_mode_i和CTRL.EN的组合状态收集功能覆盖率。'''
    },
    'TP_005': {
        'id': 'TC_005',
        'name': 'test_lane_mode',
        'description': '''TC场景：测试lane模式配置
配置条件：
1. test_mode_i=1
2. CTRL.EN=1

输入激励：
1. 配置LANE_MODE=00，发送WR_CSR命令，验证pdi_i[0]/pdo_o[0]数据传输
2. 配置LANE_MODE=01，发送WR_CSR命令，验证pdi_i[3:0]/pdo_o[3:0]数据传输(高nibble优先)
3. 尝试配置LANE_MODE=10/11(非法值)，验证不改变当前配置或返回错误
4. 在帧中间切换LANE_MODE，验证不影响当前帧

期望结果：
1. LANE_MODE=00时1-bit模式正常工作，每拍收发1bit(MSB first)
2. LANE_MODE=01时4-bit模式正常工作，每拍收发4bit(高nibble优先)
3. 非法LANE_MODE值(10/11)不改变当前配置
4. lane模式切换在帧边界生效

coverage check点：
对LANE_MODE配置值(00/01/10/11)收集功能覆盖率。'''
    },
    'TP_006': {
        'id': 'TC_006',
        'name': 'test_data_interface',
        'description': '''TC场景：测试数据接口
配置条件：
1. test_mode_i=1
2. CTRL.EN=1
3. LANE_MODE=00(1-bit)和01(4-bit)分别测试

输入激励：
1. pcs_n_i拉低，开始帧传输
2. 通过pdi_i发送opcode=0x10(WR_CSR)命令数据
3. 验证pdo_oe_o在请求阶段=0，响应阶段=1
4. 通过pdo_o接收STATUS响应
5. pcs_n_i拉高结束帧

期望结果：
1. pcs_n_i帧边界正确：拉低开始帧，拉高结束帧
2. pdi_i输入在clk_i上升沿正确采样
3. pdo_o输出在响应阶段有效
4. pdo_oe_o方向控制正确：请求=0，响应=1
5. MSB first传输顺序正确

coverage check点：
对lane模式、传输方向收集功能覆盖率。'''
    },
    'TP_007': {
        'id': 'TC_007',
        'name': 'test_protocol',
        'description': '''TC场景：测试协议帧格式
配置条件：
1. test_mode_i=1
2. CTRL.EN=1
3. LANE_MODE=00/01分别测试

输入激励：
1. 发送WR_CSR(0x10)命令：opcode=0x10 + reg_addr + wdata
2. 发送RD_CSR(0x11)命令：opcode=0x11 + reg_addr
3. 发送AHB_WR32(0x20)命令：opcode=0x20 + addr + wdata
4. 发送AHB_RD32(0x21)命令：opcode=0x21 + addr
5. 验证turnaround周期固定1个clk_i

期望结果：
1. 四种opcode对应的帧格式正确处理
2. turnaround周期固定1个clk_i
3. 协议字段MSB first传输
4. STATUS=0x00成功返回

coverage check点：
对四种opcode命令、lane模式收集功能覆盖率。'''
    },
    'TP_008': {
        'id': 'TC_008',
        'name': 'test_opcode',
        'description': '''TC场景：测试opcode解析
配置条件：
1. test_mode_i=1
2. CTRL.EN=1

输入激励：
1. 发送opcode=0x10(WR_CSR)，验证CSR写命令执行
2. 发送opcode=0x11(RD_CSR)，验证CSR读命令执行
3. 发送opcode=0x20(AHB_WR32)，验证AHB写命令执行
4. 发送opcode=0x21(AHB_RD32)，验证AHB读命令执行
5. 发送非法opcode(0x00/0xFF)，验证错误处理

期望结果：
1. opcode=0x10/0x11/0x20/0x21正确解析并执行
2. 非法opcode(0x00/0xFF)返回STS_BAD_OPCODE(0x01)
3. 前端收满8bit opcode后正确确定期望帧长

coverage check点：
对opcode值(0x10/0x11/0x20/0x21/非法值)收集功能覆盖率。'''
    },
    'TP_009': {
        'id': 'TC_009',
        'name': 'test_ctrl_config',
        'description': '''TC场景：测试CTRL配置接口
配置条件：
1. test_mode_i=1
2. 初始CTRL.EN=0

输入激励：
1. 写CTRL.EN=1，验证模块使能，发送命令可执行
2. 写CTRL.EN=0，验证模块禁用，发送命令返回STS_DISABLED(0x04)
3. 写CTRL.LANE_MODE=00，验证1-bit模式生效
4. 写CTRL.LANE_MODE=01，验证4-bit模式生效
5. 写CTRL.SOFT_RST=1，验证协议上下文清零

期望结果：
1. CTRL.EN配置正确生效
2. CTRL.LANE_MODE配置正确生效
3. CTRL.SOFT_RST=1触发软复位，恢复默认lane模式
4. 配置写入后立即生效

coverage check点：
对CTRL.EN、CTRL.LANE_MODE、CTRL.SOFT_RST配置值收集功能覆盖率。'''
    },
    'TP_010': {
        'id': 'TC_010',
        'name': 'test_status_query',
        'description': '''TC场景：测试状态查询
配置条件：
1. test_mode_i=1
2. CTRL.EN=1

输入激励：
1. 发送命令，在命令执行期间读取STATUS寄存器，验证BUSY位
2. 触发各类错误条件，读取LAST_ERR寄存器，验证错误码更新
3. 连续轮询STATUS和LAST_ERR，验证ATE轮询方式可行

期望结果：
1. STATUS.BUSY在命令执行期间=1，完成后=0
2. LAST_ERR在错误发生时更新为对应错误码
3. 无独立中断输出，状态通过寄存器查询获取
4. ATE轮询方式正确获取状态

coverage check点：
对STATUS和LAST_ERR寄存器值收集功能覆盖率。'''
    },
    'TP_011': {
        'id': 'TC_011',
        'name': 'test_error_handling',
        'description': '''TC场景：测试异常处理
配置条件：
1. test_mode_i=1
2. CTRL.EN=1

输入激励：
1. 发送非法opcode(0x00/0xFF)，验证返回STS_BAD_OPCODE(0x01)
2. 访问非法CSR地址，验证返回STS_BAD_REG(0x02)
3. 发送非4Byte对齐地址，验证返回STS_ALIGN_ERR(0x03)
4. CTRL.EN=0时发送命令，验证返回STS_DISABLED(0x04)
5. test_mode_i=0时发送命令，验证返回STS_NOT_IN_TEST(0x05)
6. 提前拉高pcs_n_i，验证返回STS_FRAME_ERR(0x06)
7. 模拟hresp_i=1，验证返回STS_AHB_ERR(0x07)
8. 模拟AHB超时，验证返回STS_TIMEOUT(0x08)

期望结果：
1. 各类错误正确检测
2. 在发起AHB访问前完成前置错误收敛
3. 返回确定性状态码
4. LAST_ERR更新正确

coverage check点：
对所有错误码类型收集功能覆盖率。'''
    },
    'TP_012': {
        'id': 'TC_012',
        'name': 'test_lowpower',
        'description': '''TC场景：测试低功耗特性
配置条件：
1. 模块处于空闲态(test_mode_i=0或无有效任务)

输入激励：
1. 保持pcs_n_i=1，监测rx_shift_reg和tx_shift_reg是否翻转
2. 模拟AHB等待态，监测超时计数器是否仅在WAIT_AHB状态工作
3. 设置test_mode_i=0，验证AHB输出保持IDLE(htrans_o=IDLE)
4. 切换到工作态，验证关键寄存器翻转

期望结果：
1. 空闲态无效翻转最小化
2. 超时计数器仅在工作状态更新
3. test_mode_i=0时AHB输出保持IDLE
4. 低功耗设计目标满足

coverage check点：
对工作状态和空闲状态的寄存器翻转次数收集功能覆盖率。'''
    },
    'TP_013': {
        'id': 'TC_013',
        'name': 'test_performance',
        'description': '''TC场景：测试性能指标
配置条件：
1. test_mode_i=1
2. CTRL.EN=1
3. clk_i=100MHz

输入激励：
1. 在1-bit模式(LANE_MODE=00)下连续执行AHB_RD32/AHB_WR32命令
2. 在4-bit模式(LANE_MODE=01)下执行相同测试
3. 测量端到端延时(从pcs_n_i拉低到pdo_oe_o拉高)
4. 计算吞吐率

期望结果：
1. 目标频率100MHz稳定工作
2. 端到端延时约23~24个clk_i周期
3. 1-bit模式原始链路线速12.5MB/s
4. 4-bit模式原始链路线速50MB/s
5. AHB访问32bit字访问，地址4Byte对齐

coverage check点：
对延时范围、吞吐率收集功能覆盖率。'''
    },
    'TP_014': {
        'id': 'TC_014',
        'name': 'test_dfx',
        'description': '''TC场景：测试DFX调试可观测点
配置条件：
1. 仿真环境
2. test_mode_i=1
3. CTRL.EN=1

输入激励：
1. 执行多种命令序列
2. 采集状态机状态、opcode、addr、wdata、rdata、status_code
3. 采集rx/tx计数、pcs_n_i、pdi_i、pdo_o、pdo_oe_o
4. 验证内部调试可观测点可访问

期望结果：
1. 状态机状态可观测
2. 关键寄存器(opcode、addr、wdata、rdata、status_code)可观测
3. 计数器(rx_count、tx_count)可观测
4. 接口信号(pcs_n_i、pdi_i、pdo_o、pdo_oe_o)可观测

coverage check点：
直接用例覆盖，不收功能覆盖率。'''
    },
    'TP_015': {
        'id': 'TC_015',
        'name': 'test_memory_map',
        'description': '''TC场景：测试memory map
配置条件：
1. test_mode_i=1
2. CTRL.EN=1

输入激励：
1. 通过WR_CSR/RD_CSR访问模块内部CSR(VERSION/CTRL/STATUS/LAST_ERR)
2. 发送AHB_WR32/AHB_RD32命令访问SoC内部memory-mapped地址
3. 验证模块自身CSR不进入SoC AHB memory map
4. 验证AHB Master访问采用32bit地址空间和word访问粒度

期望结果：
1. CSR通过协议命令访问正常
2. AHB Master访问正确
3. 模块CSR不进入SoC memory map
4. AHB访问32bit地址空间，word访问粒度

coverage check点：
对CSR地址范围、AHB地址范围收集功能覆盖率。'''
    },
    'TP_016': {
        'id': 'TC_016',
        'name': 'test_ahb_interface',
        'description': '''TC场景：测试AHB-Lite总线接口
配置条件：
1. test_mode_i=1
2. CTRL.EN=1
3. AHB总线空闲

输入激励：
1. 发送AHB_WR32(opcode=0x20)命令，验证haddr_o/hwrite_o/htrans_o/hsize_o/hburst_o/hwdata_o输出正确
2. 发送AHB_RD32(opcode=0x21)命令，验证hrdata_i/hready_i/hresp_i输入正确处理
3. 验证htrans_o仅输出IDLE(2'b00)和NONSEQ(2'b10)
4. 验证hsize_o固定为WORD(3'b010)
5. 验证hburst_o固定为SINGLE(3'b000)

期望结果：
1. AHB-Lite Master接口符合规范
2. 单次读写事务正确
3. htrans_o仅IDLE/NONSEQ
4. hsize_o固定WORD，hburst_o固定SINGLE
5. 信号时序满足AHB协议要求

coverage check点：
对AHB命令类型、地址范围收集功能覆盖率。'''
    },
}


def generate_rtm(input_path, output_path):
    """Generate new RTM file with Checker and Testcase filled."""
    # Load source workbook
    wb = openpyxl.load_workbook(input_path)

    # Get FL-TP sheet
    if 'FL-TP' not in wb.sheetnames:
        print("Error: FL-TP sheet not found")
        return False

    fl_tp_ws = wb['FL-TP']

    # Get Checker List sheet
    if 'Checker List' not in wb.sheetnames:
        print("Error: Checker List sheet not found")
        return False

    checker_ws = wb['Checker List']

    # Get DV Testcase List sheet
    if 'DV Testcase List' not in wb.sheetnames:
        print("Error: DV Testcase List sheet not found")
        return False

    tc_ws = wb['DV Testcase List']

    # Find TP entries in FL-TP and generate Checker/Testcase
    checker_row = 3  # Start from row 3 (after headers)
    tc_row = 3

    for row in range(3, fl_tp_ws.max_row + 1):
        tp_id = fl_tp_ws.cell(row=row, column=3).value
        if not tp_id or tp_id.startswith('填写'):
            continue

        # Clean tp_id (might have multiple TPs separated by newline)
        tp_ids = str(tp_id).split('\n')

        for single_tp_id in tp_ids:
            single_tp_id = single_tp_id.strip()
            if single_tp_id in CHECKERS:
                checker = CHECKERS[single_tp_id]
                tc = TESTCASES.get(single_tp_id, {})

                # Add Checker to Checker List sheet
                checker_ws.cell(row=checker_row, column=1, value=checker['id'])
                checker_ws.cell(row=checker_row, column=2, value=checker['name'])
                checker_ws.cell(row=checker_row, column=3, value=checker['description'])

                # Add Testcase to DV Testcase List sheet
                if tc:
                    tc_ws.cell(row=tc_row, column=1, value=tc['id'])
                    tc_ws.cell(row=tc_row, column=2, value=tc['name'])
                    tc_ws.cell(row=tc_row, column=3, value=tc['description'])

                # Update FL-TP links
                fl_tp_ws.cell(row=row, column=5, value=checker['id'])
                if tc:
                    fl_tp_ws.cell(row=row, column=6, value=tc['id'])

                checker_row += 1
                tc_row += 1

    # Save new workbook
    wb.save(output_path)
    print(f"RTM file generated: {output_path}")
    return True


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python generate_rtm.py <input_rtm> <output_rtm>")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2]

    generate_rtm(input_path, output_path)
