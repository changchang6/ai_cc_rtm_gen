#!/usr/bin/env python3
"""
RTM Generator Utilities
Helper functions for reading and generating RTM Excel files.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy


def read_rtm_structure(filepath):
    """Read RTM file and return structure summary."""
    wb = openpyxl.load_workbook(filepath)

    result = {
        'sheets': wb.sheetnames,
        'dr_fl': [],
        'fl_tp': [],
        'checkers': [],
        'testcases': []
    }

    # Read DR-FL sheet
    if 'DR-FL' in wb.sheetnames:
        ws = wb['DR-FL']
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] or row[2]:  # DR编号 or FL编号
                result['dr_fl'].append({
                    'dr_id': row[0],
                    'category': row[1],
                    'fl_id': row[2],
                    'description': row[3],
                    'tp_id': row[4]
                })

    # Read FL-TP sheet
    if 'FL-TP' in wb.sheetnames:
        ws = wb['FL-TP']
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[2]:  # TP编号
                result['fl_tp'].append({
                    'fl_id': row[0],
                    'tp_category': row[1],
                    'tp_id': row[2],
                    'tp_description': row[3],
                    'checker_id': row[4],
                    'testcase_id': row[5]
                })

    # Read existing Checkers
    if 'Checker List' in wb.sheetnames:
        ws = wb['Checker List']
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0]:  # CHK编号
                result['checkers'].append({
                    'chk_id': row[0],
                    'chk_name': row[1],
                    'chk_description': row[2]
                })

    # Read existing Testcases
    if 'DV Testcase List' in wb.sheetnames:
        ws = wb['DV Testcase List']
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0]:  # TC编号
                result['testcases'].append({
                    'tc_id': row[0],
                    'tc_name': row[1],
                    'tc_description': row[2]
                })

    wb.close()
    return result


def add_checker_to_rtm(wb, chk_id, chk_name, chk_description, note=None):
    """Add a checker entry to the RTM workbook."""
    ws = wb['Checker List']

    # Find next available row
    next_row = 3
    for row in range(3, ws.max_row + 1):
        if ws.cell(row=row, column=1).value is None or ws.cell(row=row, column=1).value == chk_id:
            next_row = row
            break
        next_row = row + 1

    ws.cell(row=next_row, column=1, value=chk_id)
    ws.cell(row=next_row, column=2, value=chk_name)
    ws.cell(row=next_row, column=3, value=chk_description)
    ws.cell(row=next_row, column=4, value=note)


def add_testcase_to_rtm(wb, tc_id, tc_name, tc_description, note=None):
    """Add a testcase entry to the RTM workbook."""
    ws = wb['DV Testcase List']

    # Find next available row
    next_row = 3
    for row in range(3, ws.max_row + 1):
        if ws.cell(row=row, column=1).value is None or ws.cell(row=row, column=1).value == tc_id:
            next_row = row
            break
        next_row = row + 1

    ws.cell(row=next_row, column=1, value=tc_id)
    ws.cell(row=next_row, column=2, value=tc_name)
    ws.cell(row=next_row, column=3, value=tc_description)
    ws.cell(row=next_row, column=4, value=note)


def link_tp_to_checker_testcase(wb, tp_id, checker_id, testcase_id):
    """Link a test point to checker and testcase in FL-TP sheet."""
    ws = wb['FL-TP']

    for row in range(3, ws.max_row + 1):
        if ws.cell(row=row, column=3).value == tp_id:
            ws.cell(row=row, column=5, value=checker_id)
            ws.cell(row=row, column=6, value=testcase_id)
            return True
    return False


def save_rtm(wb, filepath):
    """Save RTM workbook to file."""
    wb.save(filepath)
    return filepath


if __name__ == '__main__':
    import sys
    import json

    if len(sys.argv) < 3:
        print("Usage: python rtm_utils.py <command> <filepath> [args...]")
        print("Commands: read, check")
        sys.exit(1)

    command = sys.argv[1]
    filepath = sys.argv[2]

    if command == 'read':
        result = read_rtm_structure(filepath)
        print(json.dumps(result, ensure_ascii=False, indent=2))
    elif command == 'check':
        result = read_rtm_structure(filepath)
        print(f"DR-FL entries: {len(result['dr_fl'])}")
        print(f"FL-TP entries: {len(result['fl_tp'])}")
        print(f"Checkers: {len(result['checkers'])}")
        print(f"Testcases: {len(result['testcases'])}")
